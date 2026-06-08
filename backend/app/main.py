import os
import csv
import io
import math
import tempfile
from datetime import datetime

import openpyxl
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.database import (
    init_db, get_all_outputs, get_output,
    create_output, update_output, delete_output, get_squads,
)
from app.seed import seed_data, REFERENCE_DATA
from app.validator import validate_output

app = FastAPI(title="CSC&E Output Validator", version="2.0")


@app.get("/health")
async def health():
    return {"status": "ok"}


@app.on_event("startup")
def startup():
    init_db()
    seed_data()


class OutputCreate(BaseModel):
    csf_outcome: str = ""
    csce_outcome: str = ""
    measure: str = ""
    output_keyword: str = ""
    output: str
    impact: str = ""
    quarter: str = ""
    year: int = 2026
    priority: str = ""
    metric_baseline: float = 0
    metric_value: float = 0
    metric_target: float = 1
    output_status: str = "Not Started"
    checkpoint_status: str = ""
    checkpoint_description: str = ""
    risk: str = ""
    issue: str = ""
    owner: str = ""
    platform: str = "Cyber Security Culture & Enablement"
    squad: str = ""
    country: str = ""
    region: str = ""
    division: str = ""
    status_notes: str = ""


class OutputUpdate(OutputCreate):
    pass


class ValidateRequest(BaseModel):
    output: str
    measure: str = ""
    impact: str = ""


@app.get("/outputs")
async def list_outputs(squad: str = None):
    return get_all_outputs(squad)


@app.get("/outputs/{output_id}")
async def read_output(output_id: int):
    item = get_output(output_id)
    if not item:
        raise HTTPException(status_code=404, detail="Output not found")
    return item


@app.post("/outputs")
async def add_output(data: OutputCreate):
    output_id = create_output(data.model_dump())
    return {"id": output_id, "message": "Output created"}


@app.put("/outputs/{output_id}")
async def edit_output(output_id: int, data: OutputUpdate):
    existing = get_output(output_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Output not found")
    update_output(output_id, data.model_dump())
    return {"message": "Output updated"}


@app.delete("/outputs/{output_id}")
async def remove_output(output_id: int):
    existing = get_output(output_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Output not found")
    delete_output(output_id)
    return {"message": "Output deleted"}


@app.post("/validate")
async def validate(req: ValidateRequest):
    result = await validate_output(req.output, req.measure, req.impact)
    return result


@app.post("/outputs/{output_id}/validate")
async def validate_and_save(output_id: int):
    item = get_output(output_id)
    if not item:
        raise HTTPException(status_code=404, detail="Output not found")

    result = await validate_output(
        item["output"],
        item.get("measure", "") or "",
        item.get("impact", "") or "",
    )

    update_output(output_id, {
        "ai_status": result.get("status", ""),
        "ai_alignment": result.get("alignment", ""),
        "ai_outcome": result.get("outcome", ""),
        "ai_business_value": result.get("business_value", ""),
        "ai_output_quality": result.get("output_quality", ""),
        "ai_comment": result.get("comment", ""),
        "ai_suggestion": result.get("suggestion", ""),
        "ai_validated_on": datetime.now().isoformat(),
    })

    return result


@app.get("/reference")
async def reference_data():
    squads = get_squads()
    return {**REFERENCE_DATA, "squads": [s["name"] for s in squads]}


@app.get("/export")
async def export_csv(squad: str = None):
    rows = get_all_outputs(squad)
    if not rows:
        raise HTTPException(status_code=404, detail="No data to export")

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=rows[0].keys())
    writer.writeheader()
    writer.writerows(rows)

    response = StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
    )
    response.headers["Content-Disposition"] = "attachment; filename=csce_outputs.csv"
    return response


SHEET_SQUAD_MAP = {
    "Platform_Excellence": "Platform Excellence",
    "Platform Excellence": "Platform Excellence",
    "Human_Centric": "Human Centric Cyber Security",
    "Human Centric": "Human Centric Cyber Security",
    "Living_Security": "Living Security at Bayer",
    "Living Security": "Living Security at Bayer",
    "CSO_Greater China": "CSO - Greater China",
    "CSO_Greater_China": "CSO - Greater China",
    "CSO - Greater China": "CSO - Greater China",
    "CSO_Americas": "CSO - Americas",
    "CSO - Americas": "CSO - Americas",
    "CSO_EMEA": "CSO - EMEA",
    "CSO - EMEA": "CSO - EMEA",
    "CSO_APAC": "CSO - APAC",
    "CSO - APAC": "CSO - APAC",
}

SKIP_SHEETS = [
    "Setup Instructions", "Data Entry", "Reference Data", "Summary",
    "Dashboard", "Sheet1", "Sheet2", "Sheet3", "Instructions",
    "Lookup", "Lists", "Config",
]

COLUMN_MAP = {
    "csf outcome": "csf_outcome",
    "csce outcome": "csce_outcome",
    "measure": "measure",
    "output keyword": "output_keyword",
    "output": "output",
    "impact": "impact",
    "quarter": "quarter",
    "year": "year",
    "priority": "priority",
    "metric baseline": "metric_baseline",
    "metric value": "metric_value",
    "metric target": "metric_target",
    "output status": "output_status",
    "checkpoint status": "checkpoint_status",
    "checkpoint description": "checkpoint_description",
    "risk": "risk",
    "risks": "risk",
    "issue": "issue",
    "issues": "issue",
    "owner": "owner",
    "platform": "platform",
    "squad": "squad",
    "country": "country",
    "region": "region",
    "division": "division",
    "status notes": "status_notes",
    "notes": "status_notes",
    "status": "output_status",
    "keyword": "output_keyword",
}


def clean_val(val):
    if val is None:
        return ""
    if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
        return ""
    if isinstance(val, (int, float)):
        return val
    return str(val).strip()


@app.post("/import")
async def import_excel(file: UploadFile = File(...)):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
    except Exception as e:
        os.unlink(tmp_path)
        raise HTTPException(status_code=400, detail=f"Cannot read Excel file: {e}")

    total = 0
    sheets_imported = []

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        squad_name = SHEET_SQUAD_MAP.get(sheet_name)
        if not squad_name:
            for key, val in SHEET_SQUAD_MAP.items():
                if key.lower().replace("_", " ") in sheet_name.lower().replace("_", " "):
                    squad_name = val
                    break

        if not squad_name:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        header_idx = None
        for i, row in enumerate(rows):
            lower_vals = [str(v).lower().strip() if v else "" for v in row]
            if "output" in lower_vals:
                header_idx = i
                break

        if header_idx is None:
            continue

        headers = [str(v).lower().strip() if v else "" for v in rows[header_idx]]
        col_indices = {}
        for i, h in enumerate(headers):
            if h in COLUMN_MAP and COLUMN_MAP[h] is not None:
                db_col = COLUMN_MAP[h]
                if db_col not in col_indices:
                    col_indices[db_col] = i

        if "output" not in col_indices:
            continue

        count = 0
        for row in rows[header_idx + 1:]:
            output_idx = col_indices["output"]
            output_val = row[output_idx] if output_idx < len(row) else None

            if not output_val or str(output_val).strip() == "":
                continue

            record = {"squad": squad_name}
            for db_col, ci in col_indices.items():
                if ci < len(row):
                    record[db_col] = clean_val(row[ci])

            record["squad"] = squad_name
            record.setdefault("platform", "Cyber Security Culture & Enablement")

            if "year" in record:
                try:
                    record["year"] = int(float(str(record["year"]))) if record["year"] else 2026
                except (ValueError, TypeError):
                    record["year"] = 2026

            for nf in ["metric_baseline", "metric_value", "metric_target"]:
                if nf in record:
                    try:
                        record[nf] = float(str(record[nf])) if record[nf] != "" else 0
                    except (ValueError, TypeError):
                        record[nf] = 0

            try:
                create_output(record)
                count += 1
            except Exception:
                continue

        total += count
        sheets_imported.append({"sheet": sheet_name, "squad": squad_name, "count": count})

    wb.close()
    os.unlink(tmp_path)

    return {
        "message": f"Import complete: {total} outputs imported",
        "total": total,
        "sheets": sheets_imported,
    }
