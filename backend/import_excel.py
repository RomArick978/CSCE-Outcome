"""
CSC&E Output Validator — Excel Import Script
Imports data from "Outcome Tracking Automated 2.0.xlsx" into SQLite database.

Usage:
    python import_excel.py
    python import_excel.py "path/to/your/file.xlsx"
"""

import sys
import os
import sqlite3
import math

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# --- Configuration ---

DB_PATH = os.path.join(os.path.dirname(__file__), "data", "csce.db")

# Map sheet names to squad names
SHEET_SQUAD_MAP = {
    "Platform_Excellence": "Platform Excellence",
    "Platform Excellence": "Platform Excellence",
    "Human_Centric": "Human Centric Cyber Security",
    "Human Centric": "Human Centric Cyber Security",
    "Living_Security": "Living Security at Bayer",
    "Living Security": "Living Security at Bayer",
    "CSO_Greater China": "CSO - Greater China",
    "CSO_Greater_China": "CSO - Greater China",
    "CSO - Greater China": "CSO - Greater China",
    "CSO_Americas": "CSO - Americas",
    "CSO - Americas": "CSO - Americas",
    "CSO_EMEA": "CSO - EMEA",
    "CSO - EMEA": "CSO - EMEA",
    "CSO_APAC": "CSO - APAC",
    "CSO - APAC": "CSO - APAC",
}

# Sheets to SKIP (not data sheets)
SKIP_SHEETS = [
    "Setup Instructions",
    "Data Entry",
    "Reference Data",
    "Summary",
    "Dashboard",
    "Sheet1",
    "Sheet2",
    "Sheet3",
    "Instructions",
    "Lookup",
    "Lists",
    "Config",
]

# Map Excel column headers to DB column names
COLUMN_MAP = {
    "csf outcome": "csf_outcome",
    "csce outcome": "csce_outcome",
    "measure": "measure",
    "output keyword": "output_keyword",
    "output id": None,  # skip
    "output": "output",
    "impact": "impact",
    "quarter": "quarter",
    "year": "year",
    "priority": "priority",
    "metric baseline": "metric_baseline",
    "metric value": "metric_value",
    "metric target": "metric_target",
    "output status": "output_status",
    "checkpoint status": "checkpoint_status",
    "checkpoint description": "checkpoint_description",
    "risk": "risk",
    "risks": "risk",
    "issue": "issue",
    "issues": "issue",
    "owner": "owner",
    "platform": "platform",
    "squad": "squad",
    "country": "country",
    "region": "region",
    "division": "division",
    "status notes": "status_notes",
    "notes": "status_notes",
    "status": "output_status",
    "keyword": "output_keyword",
}


def clean_value(val):
    """Clean a cell value for database insertion."""
    if val is None:
        return ""
    if isinstance(val, float) and (math.isnan(val) or math.isinf(val)):
        return ""
    if isinstance(val, (int, float)):
        return val
    return str(val).strip()


def import_sheet(ws, sheet_name, squad_name, cursor):
    """Import a single worksheet into the database."""
    rows = list(ws.iter_rows(values_only=False))
    if len(rows) < 2:
        print(f"  Skipping {sheet_name}: no data rows")
        return 0

    # Find header row (first row with data)
    header_row = None
    for i, row in enumerate(rows):
        values = [cell.value for cell in row]
        # Check if this looks like a header row
        lower_vals = [str(v).lower().strip() if v else "" for v in values]
        if "output" in lower_vals:
            header_row = i
            break

    if header_row is None:
        print(f"  Skipping {sheet_name}: no 'Output' column found in headers")
        return 0

    # Build column index map
    headers = [str(cell.value).lower().strip() if cell.value else "" for cell in rows[header_row]]
    col_indices = {}
    for i, h in enumerate(headers):
        if h in COLUMN_MAP and COLUMN_MAP[h] is not None:
            db_col = COLUMN_MAP[h]
            if db_col not in col_indices:  # first match wins
                col_indices[db_col] = i

    if "output" not in col_indices:
        print(f"  Skipping {sheet_name}: 'output' column not mapped")
        return 0

    # Import data rows
    count = 0
    for row in rows[header_row + 1:]:
        values = [cell.value for cell in row]

        # Get output value
        output_idx = col_indices["output"]
        output_val = values[output_idx] if output_idx < len(values) else None

        # Skip empty rows
        if not output_val or str(output_val).strip() == "":
            continue

        # Build record
        record = {"squad": squad_name}
        for db_col, col_idx in col_indices.items():
            if col_idx < len(values):
                record[db_col] = clean_value(values[col_idx])

        # Override squad with sheet-based squad name
        record["squad"] = squad_name

        # Ensure required fields
        record.setdefault("output", "")
        record.setdefault("platform", "Cyber Security Culture & Enablement")

        # Handle year as integer
        if "year" in record:
            try:
                record["year"] = int(float(str(record["year"]))) if record["year"] else 2026
            except (ValueError, TypeError):
                record["year"] = 2026

        # Handle numeric fields
        for num_field in ["metric_baseline", "metric_value", "metric_target"]:
            if num_field in record:
                try:
                    record[num_field] = float(str(record[num_field])) if record[num_field] != "" else 0
                except (ValueError, TypeError):
                    record[num_field] = 0

        # Insert into database
        columns = ", ".join(record.keys())
        placeholders = ", ".join(["?" for _ in record])
        try:
            cursor.execute(
                f"INSERT INTO outputs ({columns}) VALUES ({placeholders})",
                list(record.values())
            )
            count += 1
        except Exception as e:
            print(f"  Error inserting row: {e}")
            continue

    return count


def main():
    # Determine Excel file path
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        # Try common names
        candidates = [
            "Outcome Tracking Automated 2.0.xlsx",
            "Outcome_Tracking_Automated_2.0.xlsx",
            "outcome_tracking.xlsx",
        ]
        excel_path = None
        for c in candidates:
            if os.path.exists(c):
                excel_path = c
                break

        if not excel_path:
            print("ERROR: No Excel file found.")
            print("Usage: python import_excel.py \"path/to/your/file.xlsx\"")
            print()
            print("Or place your file in this directory with one of these names:")
            for c in candidates:
                print(f"  - {c}")
            sys.exit(1)

    print(f"=" * 60)
    print(f"CSC&E Output Validator — Excel Import")
    print(f"=" * 60)
    print(f"File: {excel_path}")
    print()

    # Open Excel file
    print("Opening Excel file...")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    print(f"Sheets found: {wb.sheetnames}")
    print()

    # Ensure database exists
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # Check if outputs table exists
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='outputs'")
    if not cursor.fetchone():
        print("ERROR: Database table 'outputs' not found.")
        print("Please start the app first (uvicorn app.main:app) to initialize the database.")
        conn.close()
        sys.exit(1)

    # Process each sheet
    total = 0
    sheets_processed = 0

    for sheet_name in wb.sheetnames:
        # Skip non-data sheets
        if sheet_name in SKIP_SHEETS:
            print(f"Skipping sheet: {sheet_name} (not a data sheet)")
            continue

        # Find squad mapping
        squad_name = SHEET_SQUAD_MAP.get(sheet_name)

        if not squad_name:
            # Try partial match
            for key, val in SHEET_SQUAD_MAP.items():
                if key.lower().replace("_", " ").replace("-", " ") in sheet_name.lower().replace("_", " ").replace("-", " "):
                    squad_name = val
                    break

        if not squad_name:
            print(f"Skipping sheet: {sheet_name} (no squad mapping found)")
            continue

        print(f"Importing: {sheet_name} -> {squad_name}")
        ws = wb[sheet_name]
        count = import_sheet(ws, sheet_name, squad_name, cursor)
        print(f"  -> {count} outputs imported")
        total += count
        sheets_processed += 1

    # Commit and close
    conn.commit()
    conn.close()
    wb.close()

    print()
    print(f"=" * 60)
    print(f"IMPORT COMPLETE")
    print(f"=" * 60)
    print(f"Sheets processed: {sheets_processed}")
    print(f"Total outputs imported: {total}")
    print()
    print(f"Start the app to see your data:")
    print(f"  uvicorn app.main:app --reload --port 8000")
    print(f"  Open: http://localhost:8000")


if __name__ == "__main__":
    main()
